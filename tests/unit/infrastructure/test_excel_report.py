"""
Tests unitarios para app/infrastructure/excel/report.py.

Cubre ExcelReport.__init__() y get_xlsx() — actualmente con 12% de cobertura.
"""

import io
import os

import pytest
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helpers: tabla de ejemplo para los tests
# ---------------------------------------------------------------------------

TABLE_DEFAULTS_SINGLE = [
    {
        "title": "Hoja1",
        "subtitle": ["Informe de", "generado el", "a las"],
        "datetime_format": ["%H:%M:%S", "%d/%m/%Y"],
        "header": [
            {"header_text": "Nombre", "mapped_to": "name"},
            {"header_text": "Apellidos", "mapped_to": "surname"},
        ],
    }
]

TABLE_DEFAULTS_TWO = [
    {
        "title": "HojaA",
        "subtitle": ["Informe", "fecha", "hora"],
        "datetime_format": ["%H:%M:%S", "%d/%m/%Y"],
        "header": [
            {"header_text": "Columna1", "mapped_to": "col1"},
        ],
    },
    {
        "title": "HojaB",
        "subtitle": ["Otro informe", "fecha", "hora"],
        "datetime_format": ["%H:%M:%S", "%d/%m/%Y"],
        "header": [
            {"header_text": "Columna2", "mapped_to": "col2"},
        ],
    },
]

COMPANY_DATA = {"name": "ANPA Test", "vat_number": "B12345678", "phone": "600000000", "mail": "test@anpa.com"}

EVENTS_DATA_SINGLE = [
    [
        {"name": "Pedro", "surname": "García"},
        {"name": "Ana", "surname": "López"},
    ]
]

EVENTS_DATA_EMPTY = [[]]

EVENTS_DATA_TWO_TABLES = [
    [{"col1": "fila1"}],
    [{"col2": "fila2"}],
]


# ---------------------------------------------------------------------------
# __init__
# ---------------------------------------------------------------------------


class TestExcelReportInit:
    def test_init_creates_report_path_directory(self, tmp_path):
        """__init__() crea el directorio report_path si no existe."""
        from app.infrastructure.excel.report import ExcelReport

        report_path = str(tmp_path / "reports")
        tmp_dir = str(tmp_path / "tmp")

        assert not os.path.exists(report_path)
        ExcelReport(report_path=report_path, tmp_path=tmp_dir)
        assert os.path.exists(report_path)

    def test_init_creates_tmp_path_directory(self, tmp_path):
        """__init__() crea el directorio tmp_path si no existe."""
        from app.infrastructure.excel.report import ExcelReport

        report_path = str(tmp_path / "reports")
        tmp_dir = str(tmp_path / "tmp")

        assert not os.path.exists(tmp_dir)
        ExcelReport(report_path=report_path, tmp_path=tmp_dir)
        assert os.path.exists(tmp_dir)

    def test_init_does_not_fail_if_directories_exist(self, tmp_path):
        """__init__() no falla si los directorios ya existen."""
        from app.infrastructure.excel.report import ExcelReport

        report_path = str(tmp_path / "reports")
        tmp_dir = str(tmp_path / "tmp")
        os.makedirs(report_path)
        os.makedirs(tmp_dir)

        # No debe lanzar ninguna excepción
        ExcelReport(report_path=report_path, tmp_path=tmp_dir)


# ---------------------------------------------------------------------------
# get_xlsx() — estructura del resultado
# ---------------------------------------------------------------------------


class TestGetXlsxResult:
    def test_get_xlsx_returns_dict_with_file_data_and_filename(self, tmp_path):
        """get_xlsx() devuelve un dict con claves 'file_data' y 'filename'."""
        from app.infrastructure.excel.report import ExcelReport

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, TABLE_DEFAULTS_SINGLE, EVENTS_DATA_SINGLE)

        assert "file_data" in result
        assert "filename" in result

    def test_get_xlsx_filename_ends_with_xlsx(self, tmp_path):
        """El filename devuelto termina en .xlsx."""
        from app.infrastructure.excel.report import ExcelReport

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, TABLE_DEFAULTS_SINGLE, EVENTS_DATA_SINGLE)

        assert result["filename"].endswith(".xlsx")

    def test_get_xlsx_file_data_is_bytesio(self, tmp_path):
        """file_data es un objeto BytesIO."""
        from app.infrastructure.excel.report import ExcelReport

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, TABLE_DEFAULTS_SINGLE, EVENTS_DATA_SINGLE)

        assert isinstance(result["file_data"], io.BytesIO)

    def test_get_xlsx_file_data_is_non_empty(self, tmp_path):
        """file_data tiene contenido (no está vacío)."""
        from app.infrastructure.excel.report import ExcelReport

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, TABLE_DEFAULTS_SINGLE, EVENTS_DATA_SINGLE)

        data = result["file_data"]
        data.seek(0, 2)  # Seek to end
        size = data.tell()
        assert size > 0


# ---------------------------------------------------------------------------
# get_xlsx() — contenido del Excel
# ---------------------------------------------------------------------------


class TestGetXlsxContent:
    def _load_wb(self, tmp_path, table_defaults=None, events_data=None):
        """Helper: genera el xlsx y lo carga con openpyxl."""
        from app.infrastructure.excel.report import ExcelReport

        if table_defaults is None:
            table_defaults = TABLE_DEFAULTS_SINGLE
        if events_data is None:
            events_data = EVENTS_DATA_SINGLE

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, table_defaults, events_data)
        result["file_data"].seek(0)
        return load_workbook(result["file_data"])

    def test_get_xlsx_sheet_has_correct_title(self, tmp_path):
        """La hoja tiene el título definido en table_defaults."""
        wb = self._load_wb(tmp_path)
        assert "Hoja1" in wb.sheetnames

    def test_get_xlsx_row1_contains_company_name(self, tmp_path):
        """La fila 1 contiene el nombre de la empresa."""
        wb = self._load_wb(tmp_path)
        ws = wb["Hoja1"]
        assert ws.cell(row=1, column=1).value == COMPANY_DATA["name"]

    def test_get_xlsx_row3_contains_headers(self, tmp_path):
        """La fila 3 contiene los textos de las cabeceras de columna."""
        wb = self._load_wb(tmp_path)
        ws = wb["Hoja1"]
        header_row_values = [ws.cell(row=3, column=i).value for i in range(1, 3)]
        assert "Nombre" in header_row_values
        assert "Apellidos" in header_row_values

    def test_get_xlsx_data_starts_at_row4(self, tmp_path):
        """Los datos de eventos comienzan en la fila 4."""
        wb = self._load_wb(tmp_path)
        ws = wb["Hoja1"]
        # Row 4 debe tener al menos un valor de datos
        row4_values = [ws.cell(row=4, column=i).value for i in range(1, 3)]
        assert any(v is not None for v in row4_values)

    def test_get_xlsx_data_values_appear_in_sheet(self, tmp_path):
        """Los valores de los eventos aparecen en el cuerpo del Excel."""
        wb = self._load_wb(tmp_path)
        ws = wb["Hoja1"]
        all_values = [ws.cell(row=r, column=c).value for r in range(4, 6) for c in range(1, 3)]
        assert "Pedro" in all_values
        assert "García" in all_values

    def test_get_xlsx_empty_events_produces_sheet_with_headers_only(self, tmp_path):
        """Con events_data vacío se genera la hoja con cabeceras pero sin filas de datos."""
        wb = self._load_wb(tmp_path, events_data=EVENTS_DATA_EMPTY)
        ws = wb["Hoja1"]
        # Cabeceras en fila 3 deben existir
        header_row_values = [ws.cell(row=3, column=i).value for i in range(1, 3)]
        assert "Nombre" in header_row_values
        # Fila 4 debe estar vacía (sin datos)
        row4_values = [ws.cell(row=4, column=i).value for i in range(1, 3)]
        assert all(v is None for v in row4_values)

    def test_get_xlsx_two_tables_produce_two_sheets(self, tmp_path):
        """Con 2 table_defaults se generan 2 hojas en el workbook."""
        wb = self._load_wb(tmp_path, table_defaults=TABLE_DEFAULTS_TWO, events_data=EVENTS_DATA_TWO_TABLES)
        assert "HojaA" in wb.sheetnames
        assert "HojaB" in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_get_xlsx_can_be_opened_with_openpyxl(self, tmp_path):
        """El xlsx generado puede abrirse correctamente con openpyxl sin errores."""
        from app.infrastructure.excel.report import ExcelReport

        report = ExcelReport(report_path=str(tmp_path / "reports"), tmp_path=str(tmp_path / "tmp"))
        result = report.get_xlsx(COMPANY_DATA, TABLE_DEFAULTS_SINGLE, EVENTS_DATA_SINGLE)
        result["file_data"].seek(0)

        # No debe lanzar ninguna excepción
        wb = load_workbook(result["file_data"])
        assert wb is not None
