import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def __get_file_in_memory__(abs_filename):
    file_in_memory = None
    with open(abs_filename, "rb") as f:
        file_in_memory = io.BytesIO(f.read())
    return file_in_memory


class ExcelReport:
    def __init__(self, report_path="/tmp/excel_report", tmp_path="/tmp/excel_data"):
        self.report_path = report_path
        self.tmp_path = tmp_path

        if not os.path.exists(self.report_path):
            os.makedirs(self.report_path)

        if not os.path.exists(self.tmp_path):
            os.makedirs(self.tmp_path)

    def get_xlsx(self, company_data, table_defaults, events_data):
        generic_filename = "report_events"
        current_datetime = datetime.now()
        aux_st = f"%Y%m%d%H%M%S_{generic_filename}"

        report_name = current_datetime.strftime(aux_st)
        report_filename = f"{report_name}.xlsx"
        report_abs_filename = os.path.join(self.report_path, report_filename)
        wb = Workbook()
        counter = 0
        for item in table_defaults:
            if counter == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet(item.get("title"))
            ws.title = item.get("title")

            bold_font = Font(name="Calibri", bold=True)
            center_al = Alignment(horizontal="center")
            first_row = 1
            first_col = 1

            header_texts = item.get("header")
            if len(header_texts) > 0:
                ws.merge_cells(
                    start_row=first_row, start_column=first_col, end_row=first_row, end_column=len(header_texts)
                )
                header_company = "" + company_data.get("name")
                hc = ws.cell(column=first_col, row=first_row, value=header_company)
                hc.font = bold_font
                hc.alignment = center_al

                header_report_texts = item.get("subtitle")
                time_text = current_datetime.strftime(item.get("datetime_format")[0])
                date_text = current_datetime.strftime(item.get("datetime_format")[1])
                ws.merge_cells(
                    start_row=first_row + 1, start_column=first_col, end_row=first_row + 1, end_column=len(header_texts)
                )
                header_report_info = f"{header_report_texts[0]} {ws.title} {header_report_texts[1]} {time_text} {header_report_texts[2]} {date_text}"
                hc = ws.cell(column=first_col, row=first_row + 1, value=header_report_info)
                hc.alignment = center_al

                orders = {}
                for i in range(len(header_texts)):
                    hc = ws.cell(column=first_col + i, row=first_row + 2, value=header_texts[i]["header_text"])
                    hc.font = bold_font
                    hc.alignment = center_al
                    orders[header_texts[i]["header_text"]] = first_col + i

                for i in range(len(events_data[counter])):
                    data = events_data[counter][i]

                    for key in data.keys():
                        text = ""
                        for k in header_texts:
                            if k["mapped_to"] == key:
                                text = k["header_text"]
                        if text != "" and data[key] is not None:
                            ws.cell(column=orders[text], row=i + 4, value=data[key])

            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) * 1.1))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value

            counter = counter + 1

        wb.save(filename=report_abs_filename)

        file_in_memory = __get_file_in_memory__(report_abs_filename)
        filename = os.path.basename(report_abs_filename)
        report_file = {"file_data": file_in_memory, "filename": filename}

        return report_file
